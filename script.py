from bs4 import BeautifulSoup
import re
from openpyxl import Workbook
from openpyxl.styles import Font
from selenium import webdriver
from bs4 import BeautifulSoup

driver = webdriver.Chrome()
driver.get('https://www.imdb.com/chart/toptv')
html = driver.page_source
driver.quit()

wb = Workbook()
ws = wb.active
ws.title = "IMDB top movies"

headers = ["Name", "Rating"]
ws.append(headers)
ws['A1'].font = ws['B1'].font = ws['C1'].font = Font(bold=True)

ws.column_dimensions['A'].width = ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 40


soup = BeautifulSoup(html, 'html.parser')
els = soup.find_all('li', class_='ipc-metadata-list-summary-item')
for el in els:
  name = re.sub(r"^\d+\.\s*", "", el.select_one("h3.ipc-title__text").text)
  rating = el.select_one("span.ipc-rating-star--rating").text
  ws.append([name, rating])
wb.save("imdb-top-movies-chart.xlsx")
